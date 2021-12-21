import string
from io import BytesIO

import openpyxl

from module.MyException import NoCorrectExcelFile, ExcelNotFoundError
from openpyxl.drawing.image import Image
from PIL import Image as PImage

CM_TO_PIXEL = 37.7952755906


class Excel:
    def __init__(self, file):
        self.IMG_SIZE = {'공사사진': (11.08, 6.82), '전주번호': (3.24, 3.71)}
        self.wb = None
        self.sheet = None
        self.images = {}
        self.row = 0

        # 엑셀 파일을 불러옴
        try:
            self.wb = openpyxl.load_workbook(file)
        except FileNotFoundError:
            raise ExcelNotFoundError
        else:
            if '공정별사진대장' not in self.wb.sheetnames:
                raise NoCorrectExcelFile
            # '공정별사진대장' 시트 선택
            self.sheet = self.wb['공정별사진대장']

            # 세로 칸수를 셈
            self.row = (self.sheet.max_row - 3) // 22

            # {위치: 이미지}
            for image in self.sheet._images:
                r = image.anchor._from.row + 1
                c = string.ascii_uppercase[image.anchor._from.col]
                cell = f'{c}{r}'

                self.images[cell] = image

    def save_info(self, date, name, company, project):
        self.sheet['Q2'] = date
        self.sheet['Q3'] = name
        self.sheet['B3'] = company
        self.sheet['B2'] = project

    def findCellFrowColRow(self, col, row):
        cell_col = ['A', 'I', 'O']
        cell_row = 22 * row + 14 if cell_col[col] == 'I' else 22 * row + 6
        return f'{cell_col[col]}{cell_row}'

    def delete_image(self, col, row):
        cell = self.findCellFrowColRow(col, row)
        deleteImage = self.images[cell]
        self.sheet._images.remove(deleteImage)

    def insert_image(self, col, row, image_bytes):
        cell = self.findCellFrowColRow(col, row)
        bytes_io = self.resize_image(image_bytes)
        img = Image(bytes_io)
        size = self.IMG_SIZE['전주번호'] if cell[0] == 'I' else self.IMG_SIZE['공사사진']
        size = [cm * CM_TO_PIXEL for cm in size]
        img.width, img.height = size

        self.sheet.add_image(img, cell)

    def resize_image(self, image_bytes):
        pimg = PImage.open(image_bytes)
        pimg = pimg.resize((1200, 800))
        bytes_io = BytesIO()
        pimg.save(bytes_io, 'PNG')

        return bytes_io

    def __str__(self):
        return f'row: {self.row}, images: {self.images}'